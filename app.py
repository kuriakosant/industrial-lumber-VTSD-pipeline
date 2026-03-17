import streamlit as st
import pandas as pd
import requests
import json
import base64
import os
from io import BytesIO
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# Load environment variables
load_dotenv()

# --- CONFIGURATION ---
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "")
MODEL_NAME = "openai/gpt-4o"
TEMPLATE_PATH = "assets/ORDER-DEFAULT.xlsx"

# --- HELPER FUNCTIONS ---
def encode_image(image_bytes):
    return base64.b64encode(image_bytes).decode('utf-8')

def parse_image_with_openrouter(image_bytes):
    """Sends the image to OpenRouter and returns the structured JSON."""
    if not OPENROUTER_API_KEY:
        st.error("OpenRouter API Key is missing. Please set it in your .env file.")
        return None

    base64_image = encode_image(image_bytes)

    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }

    # Load prompt context from our rules file if it exists, otherwise use a fallback
    rules_context = ""
    rules_path = ".agent/rules/factory-logic.md"
    if os.path.exists(rules_path):
        with open(rules_path, "r") as f:
            rules_context = f.read()

    prompt = f"""
    Act as a precise data entry clerk for a wood factory. 
    You will inspect a handwritten or drawn note containing material orders.
    You must extract the Order Customer Name, Date, and a list of line-items.
    
    CRITICAL OBJECTIVE: The most important part of this process is to accurately parse the numbers (Length/ΜΗΚΟΣ, Width/ΠΛΑΤΟΣ, Quantity/ΤΕΜΑΧΙΑ) and exactly where the PVC edge banding is placed based on underlines. Getting the exact material string right is secondary to getting the dimensions and PVC rules 100% accurate.

    Follow these rules ALWAYS:
    1. STRICTLY ENGLISH: ALWAYS write ONLY in english characters (transliterate Greek to English, e.g., "ΛΕΥΚΗ" becomes "LEYKI 18MM"). NEVER output Greek text in your JSON.
    2. Grouping/Sorting: Read and sort the order exactly as if reading paragraphs from top to bottom. If there is a cluster of the same material or layout, write that entire group first before moving to the next section. Do not scramble the order.
    3. Dimensions: If an order says `80 x 76 = 2`, `80` (cm) becomes `800` (Length_mm/ΜΗΚΟΣ), and `76` (cm) becomes `760` (Width_mm/ΠΛΑΤΟΣ). Do NOT put "mm" in the final column; just output the integer (e.g., 800). `2` is the Quantity (ΤΕΜΑΧΙΑ).
    4. Edge Banding (PVC) Dashes: Customers put underlines/dashes below the numbers to indicate PVC tape for most orders:
       - 1 dash (_) below Length -> Requires PVC on 1 side. Output `2208` in `MHKOS_1` and leave `MHKOS_2` empty ("").
       - 2 dashes (__, =) below Length -> Requires PVC on BOTH sides. Output `2208` in `MHKOS_1` AND `MHKOS_2`.
       - 1 dash (_) below Width -> Output `2208` in `PLATOS_1` and leave `PLATOS_2` empty ("").
       - 2 dashes (__, =) below Width -> Output `2208` in `PLATOS_1` AND `PLATOS_2`.
       - No PVC dashes on a particular line? -> Write "OXI PVC" in the `PVC_Color` column, and leave `MHKOS_1`, `MHKOS_2`, `PLATOS_1`, and `PLATOS_2` empty (""). Do NOT leave the actual `Length_mm` and `Width_mm` empty!
    5. Columns Constraint (Material and Description): You MUST place a default material or the extracted material into the `Material` (ΥΛΙΚΟ) field for EVERY single row. NEVER leave the `Material` field empty (e.g., write "AGNOSTO" if you are truly unsure). Conversely, you MUST leave the `Description` (ΠΕΡΙΓΡΑΦΗ) field completely EMPTY ("") unless you absolutely must leave a critical warning or comment for the worker.
    6. Comments: If you have any remarks or comments about an unclear line, put them in the `Description` (ΠΕΡΙΓΡΑΦΗ) field. Do not invent custom JSON keys.

    Here are the broader factory domain rules:
    ---
    {rules_context}
    ---
    
    Output the result STRICTLY as a JSON object containing `Customer_Name`, `Date`, and `Order_Items`.
    Do NOT wrap the JSON in markdown formatting (like ```json), just output the raw JSON object.
    Ensure `MHKOS_1`, `MHKOS_2`, `PLATOS_1`, and `PLATOS_2` keys are present in every item, even if their value is an empty string "".
    """

    payload = {
        "model": MODEL_NAME,
        "messages": [
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": prompt
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ]
    }

    try:
        response = requests.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload)
        response.raise_for_status()
        
        result = response.json()
        content = result['choices'][0]['message']['content'].strip()
        
        # Clean up potential markdown formatting if the model disobeys instructions
        if content.startswith("```json"):
            content = content[7:]
        if content.endswith("```"):
            content = content[:-3]
            
        return json.loads(content)
        
    except requests.exceptions.RequestException as e:
        st.error(f"API Request failed: {e}")
        return None
    except json.JSONDecodeError as e:
        st.error(f"Failed to parse JSON response from the model. Raw response:\n{content}")
        return None
    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")
        return None

def generate_excel_from_template(customer_name, order_date, df):
    """Loads the template, injects data, and returns the byte stream."""
    if not os.path.exists(TEMPLATE_PATH):
        st.error(f"Template not found at {TEMPLATE_PATH}. Please ensure it exists.")
        return None
        
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # In openpyxl, you cannot assign a value directly to a 'MergedCell'. You must assign it to the top-left cell of the merged group.
    # Looking at the template, the data likely starts at column E (5) or F (6).
    # We will safely assign to both E and F just in case, catching the MergedCell error.
    for col in [5, 6, 7]:
        try:
            ws.cell(row=3, column=col).value = order_date
        except AttributeError:
            pass  # It's a MergedCell, ignore and try the other
            
        try:
            ws.cell(row=4, column=col).value = customer_name
        except AttributeError:
            pass

    # Define minimal styling for new rows if desired
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    # Start injecting order items at row 8
    start_row = 8
    
    # Define Column Mappings (A=1, B=2, ... K=11)
    # 1: Material, 2: Description, 3: Nera (empty), 4: Length, 5: Width, 
    # 6: Quantity, 7: PVC_Color, 8: MHKOS 1, 9: MHKOS 2, 10: PLATOS 1, 11: PLATOS 2
    
    col_mapping = {
        "Material": 1,
        "Description": 2,
        "Length_mm": 4,
        "Width_mm": 5,
        "Quantity": 6,
        "PVC_Color": 7,
        "MHKOS_1": 8,
        "MHKOS_2": 9,
        "PLATOS_1": 10,
        "PLATOS_2": 11
    }

    for index, row in df.iterrows():
        current_row = start_row + index
        
        for key, col_idx in col_mapping.items():
            val = row.get(key, "")
            # Convert to appropriate types
            if pd.isna(val) or val == "None" or str(val).strip() == "":
                val = ""
                
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = val
            
            # Apply border to match standard template
            cell.border = border_style

    output = BytesIO()
    wb.save(output)
    processed_data = output.getvalue()
    return processed_data

# --- STREAMLIT UI ---
st.set_page_config(page_title="VTSD Pipeline", layout="wide")

st.title("🏭 Factory Vision-to-Data Pipeline")
st.markdown("Upload a handwritten wood/melamine order to automatically extract it into the standard template using AI.")

# File uploader (allows camera input on mobile/tablets)
uploaded_file = st.file_uploader("Upload or take a picture of the order", type=["jpg", "jpeg", "png"])

if uploaded_file is not None:
    # Display the image side-by-side with results
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.image(uploaded_file, caption="Uploaded Image", use_column_width=True)
    
    with col2:
        if st.button("Extract Data", type="primary"):
            with st.spinner("Analyzing image and extracting symbols... This may take 15-20 seconds."):
                image_bytes = uploaded_file.getvalue()
                parsed_json = parse_image_with_openrouter(image_bytes)
                
                if parsed_json:
                    st.session_state['parsed_data'] = parsed_json
                    st.success("Analysis complete! Please review the extracted data.")

# Display and edit data below
if 'parsed_data' in st.session_state:
    st.divider()
    st.subheader("Order Validation")
    
    # Extract header info
    default_date = datetime.now().strftime("%d-%m-%Y")
    col1, col2 = st.columns(2)
    with col1:
        customer_name = st.text_input("Customer Name", value=st.session_state['parsed_data'].get("Customer_Name", ""))
    with col2:
        order_date = st.text_input("Order Date", value=st.session_state['parsed_data'].get("Date", default_date))

    st.markdown("Review and edit the line-items before generating the final Excel file. Check the dimensions and PVC tapes.")
    
    st.warning("⚠️ **DO NOT click the small download icon inside the table!** That will download a broken `.csv` file. Always use the big colored `Download Template Excel` button at the bottom of the page.")
    
    # Load items into dataframe
    df = pd.DataFrame(st.session_state['parsed_data'].get("Order_Items", []))
    
    # Ensure all columns exist even if empty
    expected_cols = ["Material", "Description", "Length_mm", "Width_mm", "Quantity", "PVC_Color", "MHKOS_1", "MHKOS_2", "PLATOS_1", "PLATOS_2"]
    for col in expected_cols:
        if col not in df.columns:
            df[col] = ""

    # Sort columns to standard order for viewing
    df = df[expected_cols]
            
    # Render editable dataframe
    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True
    )
    
    # Generate Excel button
    st.divider()
    excel_data = generate_excel_from_template(customer_name, order_date, edited_df)
    
    if excel_data:
        # Sanitize filename
        safe_name = "".join(x for x in customer_name if x.isalnum() or x in " -_").strip()
        if not safe_name:
            safe_name = "ORDER-CHECK"
            
        safe_date = "".join(x for x in order_date if x.isalnum() or x in " -_").strip()
        filename = f"{safe_name}_{safe_date}.xlsx".replace(" ", "_")
        
        st.download_button(
            label="📥 Download Template Excel",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
